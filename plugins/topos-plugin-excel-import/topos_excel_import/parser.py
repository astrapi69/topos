"""Excel parser for the Ordner-Ordnung.xlsx shape.

Four sheets, distinct semantics:

- ``"Meine Ordner"`` (29 cols): owner=SELF, type=FOLDER. Col 0 is a
  numeric external id; rows with col 0 empty either continue the
  previous container's description (when col 2 is empty and col 1
  is non-empty) or are items belonging to that container.
- ``"Ordner Eltern"`` (4 cols): owner=PARENTS, type=FOLDER. Same
  shape as the first sheet but uses only cols 0-3; no location, no
  actions.
- ``"Boxen"`` (28 cols): owner=SELF, type=BOX. Col 0 either carries a
  numeric box id or a ``"<lo> bis <hi>"`` range header that defines
  the size-group for the following boxes.

- ``"Kategorien"``: the taxonomy verbatim (path, display name, parent,
  level), so slugs and categories no item references survive.

Columns beyond the original layout (notes, category slug, owner, size
group, box priority, encoded action state) and the ``Kategorien`` sheet
are OPTIONAL - a workbook from an older version still parses with the
previous semantics. The frontend importer
(``frontend/src/excel/importWorkbook.ts``) implements the same contract;
a file written by either side imports losslessly on the other.

The parser is intentionally pure: it converts cells into in-memory
dataclasses (no DB writes). The importer module turns those records
into idempotent upserts.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import IO

import openpyxl

from .mappings import SlugifiedPath, priority_from_german, slugify_category_path

SHEET_MEINE_ORDNER = "Meine Ordner"
SHEET_ORDNER_ELTERN = "Ordner Eltern"
SHEET_BOXEN = "Boxen"
SHEET_KATEGORIEN = "Kategorien"

_RANGE_HEADER_RE = re.compile(r"^\s*(\d+)\s+bis\s+(\d+)\s*$", re.IGNORECASE)
_ACTION_SPLIT_RE = re.compile(r"\s*;\s*")
_NEGATIVE_ACTION_VALUES = {"", "keine", "nein", "no", "none"}


@dataclass
class ParsedAction:
    """One decoded action cell entry: text plus its restored state."""

    text: str
    status: str = "open"
    completed_at: datetime | None = None
    due_date: datetime | None = None


@dataclass
class ParsedItem:
    """An item row tied to the most recently seen container row.

    Translation of the Excel category cell happens during parsing so
    the importer can build the ancestor Category chain without
    re-parsing.
    """

    content: str
    external_id: int | None
    priority: str
    notes: str | None
    category_path: str | None
    category_segments: list[tuple[str, str]]
    actions: list[ParsedAction]


@dataclass
class ParsedCategory:
    """One row of the optional ``Kategorien`` sheet."""

    path: str
    display_name: str
    parent_path: str | None
    level: int


@dataclass
class ParsedContainer:
    """One container plus its child items.

    ``description_lines`` accumulates the multi-row description cells
    found beneath the container row.
    """

    external_id: int
    type: str
    owner: str
    label: str
    location: str | None
    size_group: str | None
    description_lines: list[str] = field(default_factory=list)
    items: list[ParsedItem] = field(default_factory=list)

    @property
    def description(self) -> str | None:
        if not self.description_lines:
            return None
        joined = "\n".join(line for line in self.description_lines if line)
        return joined or None


@dataclass
class ParseResult:
    """Aggregated parser output.

    ``warnings`` collects soft issues (unknown priority strings,
    unmapped category segments). The importer surfaces them in the
    HTTP response so callers can spot drift between Excel content
    and the mapping tables.
    """

    containers: list[ParsedContainer] = field(default_factory=list)
    categories: list[ParsedCategory] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _cell(row: tuple, index: int) -> object | None:
    if index >= len(row):
        return None
    return row[index]


def _cell_str(row: tuple, index: int) -> str | None:
    value = _cell(row, index)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_int(row: tuple, index: int) -> int | None:
    value = _cell(row, index)
    if value is None:
        return None
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value != int(value):
            return None
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return int(float(text))
        except ValueError:
            return None
    return None


_DONE_FLAG_RE = re.compile(r"^erledigt(?:@(.+))?$")
_ARCHIVED_FLAG_RE = re.compile(r"^archiviert(?:@(.+))?$")
_DUE_FLAG_RE = re.compile(r"^faellig:(.+)$")
_ACTION_FLAGS_RE = re.compile(r"^(.*?)\s*\[([^\]]*)\]$")


def _parse_flag_date(raw: str | None) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def _decode_action(token: str) -> ParsedAction:
    """Decode one token written by the exporter's ``_encode_action``.

    The bracket suffix counts as flags only when its content parses as
    known tokens, so ``"Regal [oben]"`` stays a plain open action
    instead of losing its brackets.
    """
    match = _ACTION_FLAGS_RE.match(token)
    if match is None:
        return ParsedAction(text=token)
    text, flag_blob = match.group(1), match.group(2)
    status = "open"
    completed_at: datetime | None = None
    due_date: datetime | None = None
    for flag in (piece.strip() for piece in flag_blob.split("|")):
        done = _DONE_FLAG_RE.match(flag)
        archived = _ARCHIVED_FLAG_RE.match(flag)
        due = _DUE_FLAG_RE.match(flag)
        if done is not None:
            status = "done"
            completed_at = _parse_flag_date(done.group(1))
        elif archived is not None:
            status = "archived"
            completed_at = _parse_flag_date(archived.group(1))
        elif due is not None:
            due_date = _parse_flag_date(due.group(1))
        else:
            # Unknown bracket content: not ours, keep the token verbatim.
            return ParsedAction(text=token)
    return ParsedAction(text=text, status=status, completed_at=completed_at, due_date=due_date)


def _split_actions(raw: str | None) -> list[ParsedAction]:
    if raw is None:
        return []
    if raw.strip().lower() in _NEGATIVE_ACTION_VALUES:
        return []
    parts = [piece.strip() for piece in _ACTION_SPLIT_RE.split(raw)]
    return [_decode_action(piece) for piece in parts if piece]


def _owner_from_cell(raw: str | None) -> str | None:
    """Read the owner column; ``None`` when absent so the sheet decides."""
    if raw is None:
        return None
    value = raw.strip().lower()
    return value if value in {"self", "parents", "shared"} else None


def _segments_from_slug_path(slug_path: str, display_cell: str | None) -> list[tuple[str, str]]:
    """Pair an explicit slug path with the German display path so each
    level keeps its display name."""
    displays = [piece.strip() for piece in (display_cell or "").split("/") if piece.strip()]
    slugs = [piece.strip() for piece in slug_path.split("/") if piece.strip()]
    return [
        (slug, displays[index] if index < len(displays) else slug)
        for index, slug in enumerate(slugs)
    ]


def _build_item(
    content: str,
    priority_cell: str | None,
    category_cell: str | None,
    notes_cell: str | None,
    action_cell: str | None,
    result: ParseResult,
    slug_path_cell: str | None = None,
    external_id: int | None = None,
) -> ParsedItem:
    priority, warning = priority_from_german(priority_cell)
    if warning:
        result.warnings.append(warning)

    # The slug column is authoritative when present: it preserves a slug
    # the German display name would not reproduce. Without it (legacy
    # workbook, hand-written file) the display path is slugified.
    if slug_path_cell:
        category_path: str | None = slug_path_cell
        category_segments = _segments_from_slug_path(slug_path_cell, category_cell)
    else:
        slug_result: SlugifiedPath | None = slugify_category_path(category_cell)
        if slug_result is not None:
            result.warnings.extend(slug_result.warnings)
        category_path = slug_result.path if slug_result else None
        category_segments = slug_result.segments if slug_result else []

    return ParsedItem(
        content=content,
        external_id=external_id,
        priority=priority,
        notes=notes_cell,
        category_path=category_path,
        category_segments=category_segments,
        actions=_split_actions(action_cell),
    )


def _parse_owner_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    owner: str,
    container_type: str,
    has_location: bool,
    has_actions: bool,
    result: ParseResult,
) -> None:
    """Parse ``Meine Ordner`` / ``Ordner Eltern``: walk top-to-bottom
    tracking the current container, attach item rows and multi-row
    description continuations."""
    current: ParsedContainer | None = None
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        external_id = _cell_int(row, 0)
        col1 = _cell_str(row, 1)
        col2 = _cell_str(row, 2)
        col3 = _cell_str(row, 3)
        col4 = _cell_str(row, 4)
        # Columns 5-6 exist on both folder sheets since the layout grew; a
        # legacy "Ordner Eltern" simply leaves them empty.
        col5 = _cell_str(row, 5)
        col6 = _cell_str(row, 6)
        notes = _cell_str(row, 7)
        slug_path = _cell_str(row, 8)
        owner_cell = _owner_from_cell(_cell_str(row, 9))
        size_group_cell = _cell_str(row, 10)
        item_number = _cell_int(row, 11)

        if external_id is not None:
            current = ParsedContainer(
                external_id=external_id,
                type=container_type,
                # The owner column wins when present; otherwise the sheet
                # decides (which is why "shared" needed a column - it
                # shares a sheet with "self").
                owner=owner_cell or owner,
                label=col1 or f"Container {external_id}",
                location=col5,
                size_group=size_group_cell,
            )
            result.containers.append(current)
            continue

        if current is None:
            # Stray data before the first container row; ignore but
            # warn so the user sees the parser dropped it.
            if col1 or col2:
                result.warnings.append(
                    f"Skipped row before first container in sheet "
                    f"{ws.title!r}: col1={col1!r} col2={col2!r}"
                )
            continue

        if col2 is not None:
            current.items.append(
                _build_item(
                    content=col2,
                    priority_cell=col3,
                    category_cell=col4,
                    notes_cell=notes,
                    action_cell=col6,
                    result=result,
                    slug_path_cell=slug_path,
                    external_id=item_number,
                )
            )
            continue

        if col1 is not None:
            # Description continuation for the current container.
            current.description_lines.append(col1)


def _parse_box_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, *, result: ParseResult) -> None:
    """Parse ``Boxen``: numeric col-0 = new box, ``"<lo> bis <hi>"``
    col-0 = size-group header, blank col-0 with col-4 = item belonging
    to the current box."""
    current_size_group: str | None = None
    current: ParsedContainer | None = None
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        col0_int = _cell_int(row, 0)
        col0_str = _cell_str(row, 0)
        col1 = _cell_str(row, 1)
        col4 = _cell_str(row, 4)
        col5 = _cell_str(row, 5)
        action_cell = _cell_str(row, 6)
        notes = _cell_str(row, 7)
        slug_path = _cell_str(row, 8)
        priority_cell = _cell_str(row, 9)
        owner_cell = _owner_from_cell(_cell_str(row, 10))
        item_number = _cell_int(row, 11)

        if col0_str is not None and col0_int is None:
            match = _RANGE_HEADER_RE.match(col0_str)
            if match is not None:
                current_size_group = f"{match.group(1)} bis {match.group(2)}"
                # Box-range description rows do not become Container records.
                continue
            # Other non-numeric col-0 strings are skipped with a warning.
            if current is None or col4 is None:
                result.warnings.append(
                    f"Skipped non-numeric row in {ws.title!r}: col0={col0_str!r}"
                )
                continue

        if col0_int is not None:
            current = ParsedContainer(
                external_id=col0_int,
                type="box",
                owner=owner_cell or "self",
                label=col1 or f"Box {col0_int}",
                location=None,
                size_group=current_size_group,
            )
            result.containers.append(current)
            continue

        if current is None:
            if col4:
                result.warnings.append(
                    f"Skipped item row before first box in {ws.title!r}: col4={col4!r}"
                )
            continue

        if col4 is not None:
            current.items.append(
                _build_item(
                    content=col4,
                    priority_cell=priority_cell,
                    category_cell=col5,
                    notes_cell=notes,
                    action_cell=action_cell,
                    result=result,
                    slug_path_cell=slug_path,
                    external_id=item_number,
                )
            )


def _parse_category_sheet(ws, *, result: ParseResult) -> None:
    """Optional ``Kategorien`` sheet: the taxonomy verbatim, so slugs,
    display names and categories no item references survive."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        path = _cell_str(row, 0)
        if path is None:
            continue
        level = _cell_int(row, 3)
        result.categories.append(
            ParsedCategory(
                path=path,
                display_name=_cell_str(row, 1) or path.rsplit("/", 1)[-1],
                parent_path=_cell_str(row, 2),
                level=level if level is not None else path.count("/"),
            )
        )


def parse_workbook(source: str | Path | IO[bytes]) -> ParseResult:
    """Parse an Ordner-Ordnung.xlsx file or bytes-like object.

    ``source`` may be a path or a file-like object. The function
    delegates to ``openpyxl.load_workbook`` with ``read_only=True``
    and ``data_only=True`` so formula cells return their cached value
    rather than the formula text.
    """
    result = ParseResult()
    wb = openpyxl.load_workbook(filename=source, read_only=True, data_only=True)
    try:
        if SHEET_MEINE_ORDNER in wb.sheetnames:
            _parse_owner_sheet(
                wb[SHEET_MEINE_ORDNER],
                owner="self",
                container_type="folder",
                has_location=True,
                has_actions=True,
                result=result,
            )
        else:
            result.warnings.append(f"Sheet {SHEET_MEINE_ORDNER!r} not found in workbook")

        if SHEET_ORDNER_ELTERN in wb.sheetnames:
            _parse_owner_sheet(
                wb[SHEET_ORDNER_ELTERN],
                owner="parents",
                container_type="folder",
                has_location=False,
                has_actions=False,
                result=result,
            )

        if SHEET_BOXEN in wb.sheetnames:
            _parse_box_sheet(wb[SHEET_BOXEN], result=result)

        if SHEET_KATEGORIEN in wb.sheetnames:
            _parse_category_sheet(wb[SHEET_KATEGORIEN], result=result)
    finally:
        wb.close()
    return result
