"""Export Topos data as an import-compatible Excel workbook."""

from __future__ import annotations

from io import BytesIO

import openpyxl
from app.models import ActionStatus, Category, Container, ContainerType, Item, Owner, Priority
from sqlalchemy.orm import Session, selectinload

from .parser import SHEET_BOXEN, SHEET_MEINE_ORDNER, SHEET_ORDNER_ELTERN

# Columns 0-6 (folder sheets) and 0-5 (box sheet) are the original
# layout; everything after was appended so an export -> import cycle
# loses nothing. Appending rather than reordering keeps workbooks written
# by older versions readable - every added column is optional on import.
OWNER_SHEET_HEADER = [
    "Nr.",
    "Ordner",
    "Inhalt",
    "Prioritaet",
    "Kategorie",
    "Ort",
    "Aktionen",
    "Notizen",
    "Kategorie-Pfad",
    "Eigentuemer",
    "Groessengruppe",
    "Eintrag-Nr.",
    "Typ",
]
BOX_SHEET_HEADER = [
    "Nr.",
    "Box",
    None,
    None,
    "Inhalt",
    "Kategorie",
    "Aktionen",
    "Notizen",
    "Kategorie-Pfad",
    "Prioritaet",
    "Eigentuemer",
    "Eintrag-Nr.",
    "Typ",
]
CATEGORY_SHEET_HEADER = ["Pfad", "Anzeigename", "Elternpfad", "Ebene"]
SHEET_KATEGORIEN = "Kategorien"

PRIORITY_LABELS = {
    Priority.VERY_HIGH: "sehr hoch",
    Priority.HIGH: "hoch",
    Priority.MEDIUM: "mittel",
    Priority.LOW: "niedrig",
    Priority.NONE: "keine",
}


def _priority_label(priority: Priority | str) -> str:
    if isinstance(priority, str):
        priority = Priority(priority)
    return PRIORITY_LABELS[priority]


def _category_display_path(categories: dict[str, Category], path: str | None) -> str | None:
    if not path:
        return None
    parts = path.split("/")
    display: list[str] = []
    for index in range(len(parts)):
        prefix = "/".join(parts[: index + 1])
        display.append(
            categories.get(prefix).display_name if prefix in categories else parts[index]
        )
    return " / ".join(display)


def _encode_action(action) -> str:
    """Encode one action so status, completion and due date survive.

    An open action without a due date stays plain text (what the format
    always looked like); anything else gets a bracket suffix, e.g.
    ``"Pruefen [erledigt@2026-01-01|faellig:2026-02-01]"``. The importer
    only reads a bracket as flags when the content parses as known
    tokens, so an action whose text ends in brackets is left alone.
    """
    flags: list[str] = []
    if action.status != ActionStatus.OPEN:
        label = "erledigt" if action.status == ActionStatus.DONE else "archiviert"
        if action.completed_at is not None:
            label = f"{label}@{action.completed_at.date().isoformat()}"
        flags.append(label)
    elif action.completed_at is not None:
        flags.append(f"erledigt@{action.completed_at.date().isoformat()}")
    if action.due_date is not None:
        flags.append(f"faellig:{action.due_date.date().isoformat()}")
    return f"{action.text} [{'|'.join(flags)}]" if flags else action.text


def _encoded_actions(container_item) -> str | None:
    """Every action of an item, encoded; ``None`` when it has none."""
    encoded = [
        _encode_action(action) for action in sorted(container_item.actions, key=lambda row: row.id)
    ]
    return "; ".join(encoded) if encoded else None


def _write_owner_sheet(
    ws,
    containers: list[Container],
    categories: dict[str, Category],
) -> None:
    ws.append(OWNER_SHEET_HEADER)
    for container in containers:
        ws.append(
            [
                container.external_id,
                container.label,
                None,
                None,
                None,
                container.location,
                None,
                None,
                None,
                container.owner.value,
                container.size_group,
                None,
                container.type.value,
            ]
        )
        if container.description:
            for line in container.description.splitlines():
                if line.strip():
                    ws.append([None, line.strip()])
        for item in sorted(container.items, key=lambda row: row.id):
            ws.append(
                [
                    None,
                    None,
                    item.content,
                    _priority_label(item.priority),
                    _category_display_path(categories, item.category_path),
                    None,
                    _encoded_actions(item),
                    item.notes,
                    item.category_path,
                    None,
                    None,
                    item.external_id,
                ]
            )


def _write_box_sheet(
    ws,
    containers: list[Container],
    categories: dict[str, Category],
) -> None:
    ws.append(BOX_SHEET_HEADER)
    current_size_group: str | None = None
    for container in containers:
        if container.size_group and container.size_group != current_size_group:
            ws.append([container.size_group])
            current_size_group = container.size_group
        ws.append(
            [
                container.external_id,
                container.label,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                container.owner.value,
                None,
                container.type.value,
            ]
        )
        if container.description:
            for line in container.description.splitlines():
                if line.strip():
                    ws.append([None, line.strip()])
        for item in sorted(container.items, key=lambda row: row.id):
            ws.append(
                [
                    None,
                    None,
                    None,
                    None,
                    item.content,
                    _category_display_path(categories, item.category_path),
                    _encoded_actions(item),
                    item.notes,
                    item.category_path,
                    _priority_label(item.priority),
                    None,
                    item.external_id,
                ]
            )


def _write_category_sheet(ws, categories: dict[str, Category]) -> None:
    """The taxonomy verbatim: keeps slugs, display names, and categories
    that no item references."""
    ws.append(CATEGORY_SHEET_HEADER)
    for category in sorted(categories.values(), key=lambda row: row.path):
        ws.append(
            [
                category.path,
                category.display_name,
                category.parent_path,
                category.level,
            ]
        )


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        max_len = max(lengths) if lengths else 0
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def export_workbook(db: Session) -> bytes:
    """Build an ``.xlsx`` workbook from the current Topos database."""
    categories = {row.path: row for row in db.query(Category).order_by(Category.path).all()}
    containers = (
        db.query(Container)
        .options(selectinload(Container.items).selectinload(Item.actions))
        .order_by(Container.external_id)
        .all()
    )
    own_folders = [
        row
        for row in containers
        if row.type == ContainerType.FOLDER and row.owner in (Owner.SELF, Owner.SHARED)
    ]
    parent_folders = [
        row for row in containers if row.type == ContainerType.FOLDER and row.owner == Owner.PARENTS
    ]
    # Everything that is not a folder shares the Boxen sheet - its owner
    # is already a column, and the appended Typ column distinguishes the
    # curated non-folder types (box, drawer, shelf, case, safe).
    boxes = [row for row in containers if row.type != ContainerType.FOLDER]

    wb = openpyxl.Workbook()
    mine = wb.active
    mine.title = SHEET_MEINE_ORDNER
    parents = wb.create_sheet(SHEET_ORDNER_ELTERN)
    box_sheet = wb.create_sheet(SHEET_BOXEN)
    category_sheet = wb.create_sheet(SHEET_KATEGORIEN)

    _write_owner_sheet(mine, own_folders, categories)
    _write_owner_sheet(parents, parent_folders, categories)
    _write_box_sheet(box_sheet, boxes, categories)
    _write_category_sheet(category_sheet, categories)
    for ws in wb.worksheets:
        _autosize(ws)
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
